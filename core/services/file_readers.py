"""
File Readers

Byte-level text extraction, one function per format. These are the fallback
readers: flat text, no structure, no page numbers. Docling handles the
structured path (see ``core/services/document_parsers``); these run for the
formats it does not cover and whenever a Docling conversion fails.

Everything here takes bytes rather than a ``File`` model instance so the same
code serves both the parser layer and any caller that already has the content
in hand.
"""

import io
import logging
import re
import zipfile
from collections import Counter
from typing import List
from xml.etree import ElementTree as ET

import fitz
import PyPDF2

logger = logging.getLogger(__name__)

# Optional spreadsheet libraries
try:
    from openpyxl import load_workbook  # type: ignore

    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd  # type: ignore

    XLRD_AVAILABLE = True
except Exception:
    XLRD_AVAILABLE = False


DOCX_NAMESPACES = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}

# Tried in order; the first that decodes cleanly wins.
TEXT_ENCODINGS = (
    "utf-8",
    "utf-8-sig",
    "latin-1",
    "cp1252",
    "iso-8859-1",
    "ascii",
)


# ============================================================================
# Public API
# ============================================================================


def read_bytes_as_text(data: bytes, filename: str) -> str:
    """Extract flat text from raw bytes, dispatching on the filename extension.

    Args:
        data: Raw file content
        filename: Original filename, used only for format detection

    Returns:
        Extracted text, or an empty string when the format carries none.
    """
    name = (filename or "").lower()

    if name.endswith(".pdf"):
        return read_pdf(data)
    if name.endswith((".txt", ".md", ".json", ".csv")):
        return decode_text(data)
    if name.endswith(".docx"):
        return read_docx(data)
    if name.endswith(".xlsx"):
        return read_xlsx(data)
    if name.endswith(".xls"):
        return read_xls(data)
    if name.endswith(".ipynb"):
        return read_ipynb(data)
    return ""


def read_ipynb(data: bytes) -> str:
    """Markdown twin of a Jupyter notebook.

    The import is inline to break a real cycle: the notebook renderer lives in
    the parser package, whose ``__init__`` reaches back into this module
    through the legacy parser.
    """
    from core.services.document_parsers.notebook_parser import notebook_markdown

    return notebook_markdown(data)


def read_pdf(data: bytes) -> str:
    """Extract native PDF text as paragraph-like blocks.

    PyMuPDF's blocks preserve enough separation to compare native text with
    Docling's structured elements. PyPDF2 remains the compatibility fallback
    for PDFs PyMuPDF cannot open.
    """
    try:
        with fitz.open(stream=data, filetype="pdf") as document:
            blocks = []
            for page in document:
                page_height = float(page.rect.height or 0)
                for block in page.get_text("blocks"):
                    text = str(block[4] or "").strip()
                    if not text:
                        continue
                    near_margin = page_height and (
                        float(block[1]) <= page_height * 0.12
                        or float(block[3]) >= page_height * 0.88
                    )
                    signature = _pdf_furniture_signature(text) if near_margin else ""
                    blocks.append((text, signature))
        repeated = Counter(signature for _, signature in blocks if signature)
        return "\n\n".join(
            text
            for text, signature in blocks
            if not signature or repeated[signature] < 2
        )
    except (RuntimeError, ValueError, TypeError) as error:
        logger.warning("PyMuPDF extraction failed; using PyPDF2: %s", error)
        reader = PyPDF2.PdfReader(io.BytesIO(data))
        return "\n\n".join(page.extract_text() or "" for page in reader.pages)


def _pdf_furniture_signature(text: str) -> str:
    """Comparable signature for repeated browser/PDF headers and footers."""
    return " ".join(re.sub(r"\d+", "#", text.casefold()).split())


def decode_text(data: bytes) -> str:
    """Decode text bytes, trying a handful of encodings before giving up."""
    for encoding in TEXT_ENCODINGS:
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def read_docx(data: bytes) -> str:
    """Extract paragraph text from a DOCX by walking ``word/document.xml``."""
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        with archive.open("word/document.xml") as document_xml:
            root = ET.fromstring(document_xml.read())

    parts = [
        element.text
        for element in root.findall(".//w:t", DOCX_NAMESPACES)
        if element.text
    ]
    return "\n".join(parts)


def read_xlsx(data: bytes) -> str:
    """Extract cell values from an XLSX workbook using openpyxl."""
    if not OPENPYXL_AVAILABLE:
        logger.warning("XLSX content not extracted: openpyxl is not installed")
        return ""

    workbook = load_workbook(filename=io.BytesIO(data), data_only=True, read_only=True)
    lines: List[str] = []
    for sheet in workbook.worksheets:
        lines.append(f"Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell) for cell in row if cell is not None]
            if values:
                lines.append("\t".join(values))
    return "\n".join(lines)


def read_xls(data: bytes) -> str:
    """Extract cell values from a legacy XLS workbook using xlrd."""
    if not XLRD_AVAILABLE:
        logger.warning("XLS content not extracted: xlrd is not installed")
        return ""

    book = xlrd.open_workbook(file_contents=data)
    lines: List[str] = []
    for sheet in book.sheets():
        lines.append(f"Sheet: {sheet.name}")
        for row_index in range(sheet.nrows):
            values = [
                str(cell)
                for cell in sheet.row_values(row_index)
                if str(cell).strip() != ""
            ]
            if values:
                lines.append("\t".join(values))
    return "\n".join(lines)
