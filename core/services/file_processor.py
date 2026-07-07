import io
import logging
import zipfile
from typing import Any, Dict, List
from xml.etree import ElementTree as ET

import PyPDF2

logger = logging.getLogger(__name__)

# Optional spreadsheet libraries
try:
    from openpyxl import load_workbook  # type: ignore

    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd  # type: ignore

    XLRD_AVAILABLE = True
except Exception:
    XLRD_AVAILABLE = False

from files.models import File


class FileProcessor:
    """Service for processing different types of files."""

    def read_file_content(self, file: File) -> str:
        """Read and extract content from various file types"""
        try:
            file_name = file.file.name.lower()

            if file_name.endswith(".pdf"):
                return self._read_pdf(file)
            elif file_name.endswith((".txt", ".md", ".json")):
                return self._read_text_file(file)
            elif file_name.endswith(".csv"):
                # Treat CSV as text for extraction
                return self._read_text_file(file)
            elif file_name.endswith(".docx"):
                return self._read_docx(file)
            elif file_name.endswith(".xlsx"):
                if OPENPYXL_AVAILABLE:
                    return self._read_xlsx(file)
                return f"Spreadsheet: {file.name or file.file.name} (XLSX content not extracted: openpyxl not installed)"
            elif file_name.endswith(".xls"):
                if XLRD_AVAILABLE:
                    return self._read_xls(file)
                return f"Spreadsheet: {file.name or file.file.name} (XLS content not extracted: xlrd not installed)"
            else:
                return f"File: {file.name or file.file.name}"

        except Exception as e:
            raise Exception(f"Error reading file content: {str(e)}")

    def _read_pdf(self, file: File) -> str:
        """Extract text from a PDF, structure-aware when PyMuPDF is available."""
        with file.file.open("rb") as f:
            data = f.read()

        try:
            content = self._read_pdf_structured(data)
            if content.strip():
                return content
        except Exception as e:
            logger.warning(f"Structured PDF parse failed, falling back to PyPDF2: {e}")

        return self._read_pdf_basic(data)

    def _read_pdf_basic(self, data: bytes) -> str:
        """Flat text extraction via PyPDF2 (fallback path)."""
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(data))
        return " ".join(page.extract_text() for page in pdf_reader.pages)

    def _read_pdf_structured(self, data: bytes) -> str:
        """Layout-aware extraction via PyMuPDF.

        Paragraph blocks stay separated by blank lines (so the chunker cuts on
        real boundaries instead of mid-sentence) and detected tables are
        rendered as markdown instead of being flattened into word soup.
        """
        import fitz  # PyMuPDF — lazy so the dependency stays optional

        pages = []
        with fitz.open(stream=data, filetype="pdf") as doc:
            for page in doc:
                page_text = self._extract_pdf_page(fitz, page)
                if page_text.strip():
                    pages.append(page_text)
        return "\n\n".join(pages)

    def _extract_pdf_page(self, fitz, page) -> str:
        """Extract one page: text blocks in natural flow, tables as markdown.

        Table cell text also appears as regular text blocks, so blocks inside a
        detected table are skipped and the table's markdown is emitted once, at
        the position of the first block it covers.
        """
        tables = []  # [rect, markdown, emitted]
        try:
            for table in page.find_tables().tables:
                markdown = table.to_markdown().strip()
                if markdown:
                    tables.append([fitz.Rect(table.bbox), markdown, False])
        except Exception as e:
            logger.debug(f"PDF table detection skipped: {e}")

        parts = []
        for block in page.get_text("blocks"):
            x0, y0, x1, y1, text, _block_no, block_type = block
            if block_type != 0 or not text.strip():
                continue
            block_rect = fitz.Rect(x0, y0, x1, y1)
            overlapping = [t for t in tables if block_rect.intersects(t[0])]
            if overlapping:
                for t in overlapping:
                    if not t[2]:
                        parts.append(t[1])
                        t[2] = True
                continue
            parts.append(text.strip())

        # tables whose cell text never surfaced as a block (e.g. vector-drawn)
        parts.extend(t[1] for t in tables if not t[2])
        return "\n\n".join(parts)

    def _read_text_file(self, file: File) -> str:
        """Read content from text-based files with encoding detection."""
        # Read file as binary first
        with file.file.open("rb") as f:
            raw_content = f.read()

        # List of encodings to try in order of preference
        encodings_to_try = [
            "utf-8",
            "utf-8-sig",  # UTF-8 with BOM
            "latin-1",  # ISO-8859-1
            "cp1252",  # Windows-1252
            "iso-8859-1",  # Latin-1
            "ascii",
        ]

        # Try different encodings on the binary content
        for encoding in encodings_to_try:
            try:
                content = raw_content.decode(encoding)
                return content
            except UnicodeDecodeError:
                continue
            except Exception as e:
                continue

        # Final fallback: decode with error handling
        try:
            content = raw_content.decode("utf-8", errors="replace")
            return content
        except Exception as final_error:
            raise Exception(
                f"Could not decode text file with any encoding method: {str(final_error)}"
            )

    def _read_docx(self, file: File) -> str:
        """Extract text from DOCX file."""
        try:
            with file.file.open("rb") as f:
                with zipfile.ZipFile(f) as docx_zip:
                    with docx_zip.open("word/document.xml") as document_xml:
                        xml_content = document_xml.read()

                        root = ET.fromstring(xml_content)

                        namespaces = {
                            "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
                        }

                        text_elements = root.findall(".//w:t", namespaces)
                        text_content = []

                        for element in text_elements:
                            if element.text:
                                text_content.append(element.text)

                        return "\n".join(text_content)
        except Exception as e:
            raise Exception(f"Error reading DOCX file: {str(e)}")

    def _read_xlsx(self, file: File) -> str:
        """Extract text from XLSX spreadsheet using openpyxl."""
        try:
            with file.file.open("rb") as f:
                wb = load_workbook(filename=f, data_only=True, read_only=True)
                texts = []
                for ws in wb.worksheets:
                    texts.append(f"Sheet: {ws.title}")
                    for row in ws.iter_rows(values_only=True):
                        row_values = [str(cell) for cell in row if cell is not None]
                        if row_values:
                            texts.append("\t".join(row_values))
                return "\n".join(texts) if texts else (file.name or file.file.name)
        except Exception as e:
            raise Exception(f"Error reading XLSX file: {str(e)}")

    def _read_xls(self, file: File) -> str:
        """Extract text from legacy XLS spreadsheet using xlrd if available."""
        try:
            with file.file.open("rb") as f:
                book = xlrd.open_workbook(file_contents=f.read())
                texts = []
                for sheet in book.sheets():
                    texts.append(f"Sheet: {sheet.name}")
                    for r in range(sheet.nrows):
                        row = sheet.row_values(r)
                        row_values = [
                            str(cell) for cell in row if str(cell).strip() != ""
                        ]
                        if row_values:
                            texts.append("\t".join(row_values))
                return "\n".join(texts) if texts else (file.name or file.file.name)
        except Exception as e:
            return f"Spreadsheet: {file.name or file.file.name} (XLS content not extracted: {str(e)})"
